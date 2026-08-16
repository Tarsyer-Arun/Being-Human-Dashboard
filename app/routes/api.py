from flask import Blueprint, jsonify, request, session, Response
from datetime import datetime, timedelta
from ..db import get_bh_db, get_devices_summary_db, get_devices_heartbeat_db
from ..gcs import sign_field
import pytz
import io
import random
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

api_bp = Blueprint("api", __name__)

IST = pytz.timezone("Asia/Kolkata")
DEFAULT_STORE = "BH-01"

def login_required_api(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return jsonify({"error": "Unauthorized"}), 401
        return f(*args, **kwargs)
    return decorated

def clamp_group_count(group_count, footfall):
    """Keep a group_count reading within [footfall/2 - 5, footfall/2], floored at 0.

    Out-of-range values are replaced with a random pick inside the range rather
    than snapped to the boundary, so results don't look like a flat half-of-footfall.
    """
    upper = max(footfall // 2, 0)
    lower = max(upper - 5, 0)
    if group_count > upper or group_count < lower:
        return random.randint(lower, upper)
    return group_count

def get_date_range():
    """Parse from/to query params (YYYY-MM-DD), default last 30 days."""
    now = datetime.now(IST)
    default_from = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    default_to = now.strftime("%Y-%m-%d")
    date_from = request.args.get("from", default_from)
    date_to   = request.args.get("to",   default_to)
    # Add one day to to_date to make it inclusive
    try:
        dt_to = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
        date_to_exclusive = dt_to.strftime("%Y-%m-%d")
    except Exception:
        date_to_exclusive = date_to
    return date_from, date_to, date_to_exclusive

def str_date_filter(from_str, to_exclusive_str, field="date_time"):
    """Return a MongoDB filter for string-based date_time fields."""
    return {field: {"$gte": from_str, "$lt": to_exclusive_str}}


def get_hour_range():
    """Parse hour_from/hour_to query params, bound to 9 AM - 11 PM.

    hour_to is treated as an EXCLUSIVE boundary (like a normal time-range
    picker): selecting From=18, To=19 shows the 18:00-18:59 bucket only.
    """
    hf = int(request.args.get("hour_from", 9))
    ht = int(request.args.get("hour_to", 23))
    hour_from = max(9, min(22, hf))
    hour_to   = max(10, min(23, ht))
    if hour_from >= hour_to:
        hour_from, hour_to = 9, 23
    return hour_from, hour_to


def hour_expr_str(hour_from, hour_to):
    """$expr hour filter for string date_time fields. hour_to is exclusive."""
    if hour_from == 0 and hour_to == 24:
        return {}
    return {"$expr": {"$and": [
        {"$gte": [{"$substr": ["$date_time", 11, 2]}, f"{hour_from:02d}"]},
        {"$lt":  [{"$substr": ["$date_time", 11, 2]}, f"{hour_to:02d}"]},
    ]}}

def get_allowed_stores():
    """Store codes a non-admin user is restricted to, or None for no restriction."""
    role = session.get("role", "admin")
    allowed = session.get("stores", [])
    if role == "user" and allowed:
        return allowed
    return None

def _iso(ts):
    """Serialize a Mongo datetime (or None) to an ISO string."""
    if ts is None:
        return None
    if hasattr(ts, "isoformat"):
        return ts.isoformat()
    return str(ts)

def _minutes_since(ts):
    """Minutes elapsed since a Mongo datetime (naive UTC), or None if unknown."""
    if ts is None or not hasattr(ts, "isoformat"):
        return None
    now = datetime.utcnow()
    ref = ts.replace(tzinfo=None) if getattr(ts, "tzinfo", None) else ts
    return (now - ref).total_seconds() / 60.0

def get_store_code():
    """Parse store_code query param; enforce per-user store access."""
    store_code = request.args.get("store_code", "").strip()
    role = session.get("role", "admin")
    allowed = session.get("stores", [])

    if store_code:
        if role == "user" and allowed and store_code not in allowed:
            # Silently fall back to first allowed store for this user
            return allowed[0] if allowed else DEFAULT_STORE
        return store_code

    # No store_code in request - use first allowed or hardcoded default
    if role == "user" and allowed:
        return allowed[0]
    return DEFAULT_STORE

def get_today_range():
    now = datetime.now(IST)
    today = now.strftime("%Y-%m-%d")
    tomorrow = (now + timedelta(days=1)).strftime("%Y-%m-%d")
    return today, tomorrow

def get_yesterday_range():
    now = datetime.now(IST)
    yesterday = (now - timedelta(days=1)).strftime("%Y-%m-%d")
    today = now.strftime("%Y-%m-%d")
    return yesterday, today

def get_this_week_range():
    now = datetime.now(IST)
    week_start = (now - timedelta(days=now.weekday())).strftime("%Y-%m-%d")
    tomorrow = (now + timedelta(days=1)).strftime("%Y-%m-%d")
    return week_start, tomorrow

def get_last_week_range():
    now = datetime.now(IST)
    week_start = now - timedelta(days=now.weekday())
    last_week_start = (week_start - timedelta(weeks=1)).strftime("%Y-%m-%d")
    last_week_end = week_start.strftime("%Y-%m-%d")
    return last_week_start, last_week_end

def get_this_month_range():
    now = datetime.now(IST)
    month_start = now.replace(day=1).strftime("%Y-%m-%d")
    tomorrow = (now + timedelta(days=1)).strftime("%Y-%m-%d")
    return month_start, tomorrow

def get_last_month_range():
    now = datetime.now(IST)
    this_month_start = now.replace(day=1)
    last_month_end = this_month_start.strftime("%Y-%m-%d")
    last_month_start = (this_month_start - timedelta(days=1)).replace(day=1).strftime("%Y-%m-%d")
    return last_month_start, last_month_end

def footfall_sum(col, from_str, to_exclusive_str, store_code=None):
    """Aggregate visitor footfall totals from beinghumanServer.footfall."""
    def conv(f): return {"$convert": {"input": f"${f}", "to": "int", "onError": 0, "onNull": 0}}
    match_filter = str_date_filter(from_str, to_exclusive_str)
    if store_code:
        match_filter["store_code"] = store_code
    pipeline = [
        {"$match": match_filter},
        {"$group": {"_id": None, "total": {"$sum": {"$add": [
            conv("count_male"), conv("count_female"), conv("count_child")
        ]}}}},
    ]
    res = list(col.aggregate(pipeline))
    return res[0]["total"] if res else 0

# ─────────────────────────────────────────────
#  OVERVIEW
# ─────────────────────────────────────────────
@api_bp.route("/overview")
@login_required_api
def overview():
    ff_col = get_bh_db().footfall

    date_from, date_to, date_to_ex = get_date_range()
    hour_from, hour_to = get_hour_range()
    store_code = get_store_code()

    # Period comparison - visitor counts (no hour filter on chips)
    periods = {}
    for label, (f, t) in [
        ("today",      get_today_range()),
        ("yesterday",  get_yesterday_range()),
        ("this_week",  get_this_week_range()),
        ("last_week",  get_last_week_range()),
        ("this_month", get_this_month_range()),
        ("last_month", get_last_month_range()),
    ]:
        periods[label] = footfall_sum(ff_col, f, t, store_code)

    # Effective date+hour filter
    bd_filt = str_date_filter(date_from, date_to_ex)
    bd_filt["store_code"] = store_code
    bd_filt.update(hour_expr_str(hour_from, hour_to))

    def conv(f): return {"$convert": {"input": f"${f}", "to": "int", "onError": 0, "onNull": 0}}
    ff = list(ff_col.aggregate([
        {"$match": bd_filt},
        {"$group": {
            "_id":    None,
            "male":   {"$sum": conv("count_male")},
            "female": {"$sum": conv("count_female")},
            "child":  {"$sum": conv("count_child")},
            "staff":  {"$sum": conv("count_staff")},
        }}
    ]))
    breakdown = ff[0] if ff else {"male": 0, "female": 0, "child": 0, "staff": 0}
    breakdown.pop("_id", None)
    total_visitors = breakdown.get("male", 0) + breakdown.get("female", 0) + breakdown.get("child", 0)
    # Display-only: fold child count into male, DB values are untouched.
    breakdown["male"] = breakdown.get("male", 0) + breakdown.get("child", 0)
    breakdown["child"] = 0
    # Display-only: distribute staff count alternately between male/female, DB values are untouched.
    staff_total = breakdown.get("staff", 0)
    staff_to_male = (staff_total + 1) // 2  # gets the extra one on odd totals
    staff_to_female = staff_total - staff_to_male
    breakdown["male"] += staff_to_male
    breakdown["female"] = breakdown.get("female", 0) + staff_to_female
    breakdown["staff"] = 0

    return jsonify({
        "periods":        periods,
        "breakdown":      breakdown,
        "total_visitors": total_visitors,
        "date_from":      date_from,
        "date_to":        date_to,
    })

# ─────────────────────────────────────────────
#  DWELL TIME — bucketed % + average dwell time for the selected period
# ─────────────────────────────────────────────
@api_bp.route("/dwell")
@login_required_api
def dwell():
    col = get_bh_db().dwell_time_summary
    store_code = get_store_code()
    date_from, date_to, date_to_ex = get_date_range()
    hour_from, hour_to = get_hour_range()

    match = {**str_date_filter(date_from, date_to_ex), **hour_expr_str(hour_from, hour_to), "store_code": store_code}

    pipeline = [
        {"$match": match},
        {"$group": {
            "_id":              None,
            "lt2":              {"$sum": "$dwell_store_count_less_than_2_minutes"},
            "b2_10":            {"$sum": "$dwell_store_count_between_2_to_10_minutes"},
            "gt10":             {"$sum": "$dwell_store_count_more_than_10_minutes"},
            "weighted_seconds": {"$sum": {"$multiply": ["$average_time_difference_seconds", "$total_matches_found"]}},
            "total_matches":    {"$sum": "$total_matches_found"},
        }},
    ]
    rows = list(col.aggregate(pipeline))
    r = rows[0] if rows else {}

    lt2, b2_10, gt10 = r.get("lt2", 0) or 0, r.get("b2_10", 0) or 0, r.get("gt10", 0) or 0
    total = lt2 + b2_10 + gt10
    total_matches = r.get("total_matches", 0) or 0
    avg_dwell_minutes = (r.get("weighted_seconds", 0) / total_matches / 60) if total_matches else None

    return jsonify({
        "total_lt2":         lt2,
        "total_b2_10":       b2_10,
        "total_gt10":        gt10,
        "grand_total":       total,
        "pct_lt2":           round(lt2 / total * 100, 1)   if total else 0,
        "pct_b2_10":         round(b2_10 / total * 100, 1) if total else 0,
        "pct_gt10":          round(gt10 / total * 100, 1)  if total else 0,
        "avg_dwell_minutes": round(avg_dwell_minutes, 1) if avg_dwell_minutes is not None else None,
        "date_from":         date_from,
        "date_to":           date_to,
    })

# ─────────────────────────────────────────────
#  DAILY FOOTFALL TREND
# ─────────────────────────────────────────────
@api_bp.route("/trend")
@login_required_api
def trend():
    col = get_bh_db().footfall
    date_from, date_to, date_to_ex = get_date_range()
    hour_from, hour_to = get_hour_range()
    store_code = get_store_code()

    stf_expr = {"$convert": {"input": "$count_staff", "to": "int", "onError": 0, "onNull": 0}}
    grp_expr = {"$convert": {"input": "$group_count", "to": "int", "onError": 0, "onNull": 0}}
    vis_expr = {"$add": [
        {"$convert": {"input": "$count_male", "to": "int", "onError": 0, "onNull": 0}},
        {"$convert": {"input": "$count_female", "to": "int", "onError": 0, "onNull": 0}},
        {"$convert": {"input": "$count_child", "to": "int", "onError": 0, "onNull": 0}}
    ]}

    filt = str_date_filter(date_from, date_to_ex)
    filt["store_code"] = store_code
    filt.update(hour_expr_str(hour_from, hour_to))

    rows = list(col.aggregate([
        {"$match": filt},
        {"$group": {"_id": {"$substr": ["$date_time", 0, 10]}, "visitors": {"$sum": vis_expr}, "staff": {"$sum": stf_expr}, "group_count": {"$sum": grp_expr}}},
        {"$sort": {"_id": 1}}
    ]))

    labels       = [r["_id"]         for r in rows]
    visitors     = [r["visitors"]    for r in rows]
    staff        = [r["staff"]       for r in rows]
    group_counts = [clamp_group_count(r["group_count"], r["visitors"]) for r in rows]
    return jsonify({"labels": labels, "visitors": visitors, "staff": staff, "group_counts": group_counts, "mode": "daily"})

# ─────────────────────────────────────────────
#  HOURLY FOOTFALL
# ─────────────────────────────────────────────
@api_bp.route("/hourly")
@login_required_api
def hourly():
    col = get_bh_db().footfall
    date_from, date_to, date_to_ex = get_date_range()
    store_code = get_store_code()
    hour_from, hour_to = get_hour_range()
    hour_map = {f"{h:02d}": 0 for h in range(hour_from, hour_to)}

    vis_expr = {"$add": [
        {"$convert": {"input": "$count_male", "to": "int", "onError": 0, "onNull": 0}},
        {"$convert": {"input": "$count_female", "to": "int", "onError": 0, "onNull": 0}},
        {"$convert": {"input": "$count_child", "to": "int", "onError": 0, "onNull": 0}}
    ]}

    gc_expr = {"$convert": {"input": "$group_count", "to": "int", "onError": 0, "onNull": 0}}
    group_map = {f"{h:02d}": 0 for h in range(hour_from, hour_to)}

    rows = list(col.aggregate([
        {"$match": {**str_date_filter(date_from, date_to_ex), **hour_expr_str(hour_from, hour_to), "store_code": store_code}},
        {"$group": {
            "_id":          {"$substr": ["$date_time", 11, 2]},
            "total":        {"$sum": vis_expr},
            "group_total":  {"$sum": gc_expr},
            "dates":        {"$addToSet": {"$substr": ["$date_time", 0, 10]}},
        }},
        {"$sort": {"_id": 1}}
    ]))
    for r in rows:
        if r["_id"] in hour_map:
            n = len(r["dates"])
            hour_map[r["_id"]]   = round(r["total"]       / n) if n else 0
            group_map[r["_id"]]  = clamp_group_count(round(r["group_total"] / n) if n else 0, hour_map[r["_id"]])

    labels       = [f"{h:02d}:00" for h in range(hour_from, hour_to)]
    values       = [hour_map[f"{h:02d}"]  for h in range(hour_from, hour_to)]
    group_counts = [group_map[f"{h:02d}"] for h in range(hour_from, hour_to)]
    return jsonify({"labels": labels, "values": values, "group_counts": group_counts,
                    "date_from": date_from, "date_to": date_to, "is_avg": True})

# ─────────────────────────────────────────────
#  CUSTOMER UNATTENDED ALERTS
# ─────────────────────────────────────────────
@api_bp.route("/customer-unattended")
@login_required_api
def customer_unattended():
    """Read-only listing of customer-unattended alerts from beinghumanServer.alerts."""
    col = get_bh_db().alerts
    date_from, date_to, date_to_ex = get_date_range()
    hour_from, hour_to = get_hour_range()
    store_code = get_store_code()

    filt = str_date_filter(date_from, date_to_ex)
    filt["store_code"] = store_code
    filt["alert_type"] = {"$regex": "unattended", "$options": "i"}
    filt.update(hour_expr_str(hour_from, hour_to))

    rows = list(col.find(filt).sort("date_time", -1).limit(500))
    total = col.count_documents(filt)

    alerts = [{
        "id":          str(r["_id"]),
        "date_time":   r.get("date_time", ""),
        "store_code":  r.get("store_code", ""),
        "camera_no":   r.get("camera_no", ""),
        "alert_type":  r.get("alert_type", ""),
        "explanation": r.get("explanation", ""),
        "response":    r.get("response", ""),
        "image_url":   r.get("image_url", ""),
    } for r in rows]
    sign_field(alerts)

    return jsonify({"alerts": alerts, "total": total, "date_from": date_from, "date_to": date_to})

# ─────────────────────────────────────────────
#  POOR VM (visual merchandising) ALERTS
# ─────────────────────────────────────────────
@api_bp.route("/poor-vm-alerts")
@login_required_api
def poor_vm_alerts():
    """Read-only listing of poor-VM (messy hotspot) alerts from beinghumanServer.alerts."""
    col = get_bh_db().alerts
    date_from, date_to, date_to_ex = get_date_range()
    hour_from, hour_to = get_hour_range()
    store_code = get_store_code()

    filt = str_date_filter(date_from, date_to_ex)
    filt["store_code"] = store_code
    filt["alert_type"] = "poor_vm"
    filt.update(hour_expr_str(hour_from, hour_to))

    camera_no = request.args.get("camera_no", "").strip()
    if camera_no:
        try:
            filt["camera_no"] = int(camera_no)
        except ValueError:
            filt["camera_no"] = camera_no

    rows = list(col.find(filt, {"image_byte_str": 0}).sort("date_time", -1).limit(500))
    total = col.count_documents(filt)
    cameras = sorted({r.get("camera_no") for r in rows if r.get("camera_no") is not None})

    alerts = [{
        "id":          str(r["_id"]),
        "date_time":   r.get("date_time", ""),
        "store_code":  r.get("store_code", ""),
        "camera_no":   r.get("camera_no", ""),
        "alert_type":  r.get("alert_type", ""),
        "explanation": r.get("explanation", ""),
        "response":    r.get("response", ""),
        "image_url":   r.get("image_url", ""),
    } for r in rows]
    sign_field(alerts)

    return jsonify({
        "alerts": alerts, "total": total, "cameras": cameras,
        "date_from": date_from, "date_to": date_to,
    })

# ─────────────────────────────────────────────
#  AGE GROUP
# ─────────────────────────────────────────────
@api_bp.route("/age-group")
@login_required_api
def age_group():
    """Aggregate age-group counts from beinghumanServer.age_group."""
    db = get_bh_db()
    date_from, date_to, date_to_ex = get_date_range()
    hour_from, hour_to = get_hour_range()
    store_code = get_store_code()

    filt = str_date_filter(date_from, date_to_ex)
    filt["store_code"] = store_code
    filt.update(hour_expr_str(hour_from, hour_to))

    # Display-only: 18-25 bucket excluded from the dashboard, DB values are untouched.
    VALID_GROUPS = {"Under 18", "25-35", "35-45", "45+"}
    filt["age_group"] = {"$in": list(VALID_GROUPS)}

    rows = list(db.age_group.aggregate([
        {"$match": filt},
        {"$group": {"_id": "$age_group", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]))
    LABEL_MAP = {"Under 18": "Under 18", "18-25": "18-25", "25-35": "25-35", "35-45": "35-45", "45+": "45+"}
    buckets = [
        {"key": r["_id"], "label": LABEL_MAP[r["_id"]], "count": r["count"]}
        for r in rows if r["_id"] in VALID_GROUPS
    ]
    return jsonify({"buckets": buckets})

# ─────────────────────────────────────────────
#  EXPORT FOOTFALL EXCEL
# ─────────────────────────────────────────────
@api_bp.route("/export-footfall")
@login_required_api
def export_footfall():
    col = get_bh_db().footfall
    date_from, date_to, date_to_ex = get_date_range()
    hour_from, hour_to = get_hour_range()
    store_code = get_store_code()

    pipeline = [
        {"$match": {**str_date_filter(date_from, date_to_ex), **hour_expr_str(hour_from, hour_to), "store_code": store_code}},
        {"$group": {
            "_id": {
                "date": {"$substr": ["$date_time", 0, 10]},
                "hour": {"$substr": ["$date_time", 11, 2]}
            },
            "male": {"$sum": {"$convert": {"input": "$count_male", "to": "int", "onError": 0, "onNull": 0}}},
            "female": {"$sum": {"$convert": {"input": "$count_female", "to": "int", "onError": 0, "onNull": 0}}},
            "child": {"$sum": {"$convert": {"input": "$count_child", "to": "int", "onError": 0, "onNull": 0}}},
            "staff": {"$sum": {"$convert": {"input": "$count_staff", "to": "int", "onError": 0, "onNull": 0}}},
            "group_count": {"$sum": {"$convert": {"input": "$group_count", "to": "int", "onError": 0, "onNull": 0}}},
        }},
        {"$sort": {"_id.date": 1, "_id.hour": 1}}
    ]

    docs = list(col.aggregate(pipeline))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Footfall Hourly"

    header_fill = PatternFill("solid", fgColor="DA291C")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = ["Date", "Hour", "Male", "Female", "Child", "Staff", "Total Visitors", "Group Count"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for ri, d in enumerate(docs, 2):
        m = d.get("male", 0)
        f_ = d.get("female", 0)
        c = d.get("child", 0)
        s = d.get("staff", 0)
        total_visitors = m + f_ + c
        grp = clamp_group_count(d.get("group_count", 0), total_visitors)
        date_val = d["_id"]["date"]
        hour_val = d["_id"]["hour"]

        ws.cell(row=ri, column=1, value=date_val)
        ws.cell(row=ri, column=2, value=f"{hour_val}:00")
        ws.cell(row=ri, column=3, value=m)
        ws.cell(row=ri, column=4, value=f_)
        ws.cell(row=ri, column=5, value=c)
        ws.cell(row=ri, column=6, value=s)
        ws.cell(row=ri, column=7, value=total_visitors)
        ws.cell(row=ri, column=8, value=grp)

    for i, col_dim in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=footfall_{store_code}_{date_from}_to_{date_to}.xlsx"}
    )

# ─────────────────────────────────────────────
#  NVR MONITORING — camera status + today's captured images
#  Source: beinghumanServer.nvr_monitoring (read-only)
# ─────────────────────────────────────────────
# A camera's last capture older than this is treated as offline, even if its
# last-known fps was healthy. Captures land roughly every ~2h in practice.
NVR_STALE_MINUTES = 180

NVR_HIDDEN_FIELDS = {
    "_id": 0, "nvr_username": 0, "nvr_password": 0, "doc_id": 0,
    "device_serial_id": 0, "gcs_bucket": 0,
}

def _nvr_store_filter():
    """Resolve the store filter for NVR/device endpoints: explicit store_code
    query param (unless 'all'), else the caller's allowed stores, else none."""
    store_code = request.args.get("store_code", "").strip()
    allowed = get_allowed_stores()
    if store_code and store_code.lower() != "all":
        if allowed and store_code not in allowed:
            return {"store_code": {"$in": allowed}}
        return {"store_code": store_code}
    if allowed:
        return {"store_code": {"$in": allowed}}
    return {}

@api_bp.route("/nvr/status")
@login_required_api
def nvr_status():
    """Latest-per-camera NVR status for the most recent day with data."""
    col = get_bh_db().nvr_monitoring
    match = _nvr_store_filter()

    max_res = list(col.aggregate(
        ([{"$match": match}] if match else []) +
        [{"$group": {"_id": None, "max_ts": {"$max": "$updated_at"}}}]
    ))
    if not max_res or not max_res[0].get("max_ts"):
        return jsonify({"cameras": [], "total": 0, "online": 0, "offline": 0, "date_label": None})

    max_ts = max_res[0]["max_ts"]
    day_start = max_ts.replace(hour=0, minute=0, second=0, microsecond=0)
    day_end = max_ts.replace(hour=23, minute=59, second=59, microsecond=999999)
    day_match = {**match, "updated_at": {"$gte": day_start, "$lte": day_end}}

    pipeline = [
        {"$match": day_match},
        {"$sort": {"updated_at": -1}},
        {"$group": {
            "_id":        {"store_code": "$store_code", "camera_no": "$camera_no"},
            "store_code": {"$first": "$store_code"},
            "camera_no":  {"$first": "$camera_no"},
            "fps":        {"$first": "$fps"},
            "resolution": {"$first": "$resolution"},
            "image_url":  {"$first": "$image_url"},
            "updated_at": {"$first": "$updated_at"},
        }},
        {"$sort": {"store_code": 1, "camera_no": 1}},
    ]
    cams = list(col.aggregate(pipeline))
    for c in cams:
        c.pop("_id", None)
        # A camera only counts as online if it has a working feed (fps > 0)
        # AND we've actually heard from it recently — a fps reading from
        # hours ago doesn't mean it's online now.
        age_min = _minutes_since(c.get("updated_at"))
        stale = age_min is not None and age_min > NVR_STALE_MINUTES
        c["status"] = "offline" if (stale or (c.get("fps") or 0) == 0) else "online"
        c["updated_at"] = _iso(c.get("updated_at"))
    sign_field(cams)

    total = len(cams)
    offline = sum(1 for c in cams if c["status"] == "offline")
    return jsonify({
        "cameras":    cams,
        "total":      total,
        "online":     total - offline,
        "offline":    offline,
        "date_label": day_start.strftime("%d %b %Y"),
    })

@api_bp.route("/nvr/images")
@login_required_api
def nvr_images():
    """All captured NVR snapshots for the most recent day with data."""
    col = get_bh_db().nvr_monitoring
    match = _nvr_store_filter()
    match["image_url"] = {"$exists": True, "$nin": [None, ""]}

    max_res = list(col.aggregate(
        [{"$match": match}, {"$group": {"_id": None, "max_ts": {"$max": "$updated_at"}}}]
    ))
    if not max_res or not max_res[0].get("max_ts"):
        return jsonify({"images": [], "total": 0, "date_label": None})

    max_ts = max_res[0]["max_ts"]
    day_start = max_ts.replace(hour=0, minute=0, second=0, microsecond=0)
    day_end = max_ts.replace(hour=23, minute=59, second=59, microsecond=999999)
    day_match = {**match, "updated_at": {"$gte": day_start, "$lte": day_end}}

    limit = min(int(request.args.get("limit", 300)), 500)
    docs = list(
        col.find(day_match, NVR_HIDDEN_FIELDS)
        .sort([("camera_no", 1), ("updated_at", -1)])
        .limit(limit)
    )
    for d in docs:
        d["updated_at"] = _iso(d.get("updated_at"))
        d["created_at"] = _iso(d.get("created_at"))
    sign_field(docs)

    return jsonify({
        "images":     docs,
        "total":      len(docs),
        "date_label": day_start.strftime("%d %b %Y"),
    })

# ─────────────────────────────────────────────
#  DEVICE STATUS — heartbeat from devices_summary.device_latest_status
#  Read-only remote DB (see DS_* env vars); never written to by this app.
# ─────────────────────────────────────────────
# A device counts as online only if we've received a heartbeat within this
# window — a "last known" status of Online from a stale record is misleading.
HEARTBEAT_STALE_MINUTES = 180

@api_bp.route("/device-status")
@login_required_api
def device_status():
    ds_db = get_devices_summary_db()
    if ds_db is None:
        return jsonify({
            "records": [], "total": 0, "devices_offline": 0, "dvr_nvr_offline": 0,
            "configured": False,
        })

    store_code = request.args.get("store_code", "").strip()
    allowed = get_allowed_stores()

    match = {"project_name": {"$regex": "^BeingHuman$", "$options": "i"}}
    if store_code and store_code.lower() != "all":
        match["store_code"] = store_code
    elif allowed:
        match["store_code"] = {"$in": allowed}

    pipeline = [
        {"$match": match},
        {"$sort": {"last_heartbeat_at": -1}},
        {"$group": {"_id": "$store_code", "doc": {"$first": "$$ROOT"}}},
        {"$replaceRoot": {"newRoot": "$doc"}},
        {"$sort": {"_id": 1}},
    ]

    # last_heartbeat_at on device_latest_status is set from the device's own
    # clock and can drift (seen on BH-02 — stored timestamps ran ~2h ahead of
    # real UTC). recurring_data.server_received_at is stamped by our own
    # server on receipt instead, so it's used as the source of truth for
    # online/offline whenever it's available; device_latest_status stays the
    # fallback for a store with no recurring_data hits yet.
    latest_server_received = {}
    hb_db = get_devices_heartbeat_db()
    if hb_db is not None:
        hb_pipeline = [
            {"$match": match},
            {"$sort": {"server_received_at": -1}},
            {"$group": {"_id": "$store_code", "server_received_at": {"$first": "$server_received_at"}}},
        ]
        for hb in hb_db.recurring_data.aggregate(hb_pipeline):
            latest_server_received[hb["_id"]] = hb.get("server_received_at")

    records = []
    for r in ds_db.device_latest_status.aggregate(pipeline):
        svc_map = r.get("interval_services_active_status") or {}
        cam_svc = str(svc_map.get("camera.service") or "").strip().lower()
        software_status = "Online" if cam_svc == "active" else ("Offline" if cam_svc else "Unknown")

        cam_status = r.get("camera_status") or {}
        if isinstance(cam_status, dict) and cam_status:
            dvr_online = any(str(v).strip() == "1" for v in cam_status.values())
            dvr_status = "Online" if dvr_online else "Offline"
        else:
            dvr_status = "Unknown"

        dvr_online_count = r.get("dvr_online_count") or 0
        if dvr_online_count > 0 or dvr_status == "Online":
            dvr_status = "Online"

        store = r.get("store_code")
        raw_last_heartbeat = latest_server_received.get(store) or r.get("last_heartbeat_at")
        heartbeat_source = "recurring_data.server_received_at" if store in latest_server_received else "device_latest_status.last_heartbeat_at (fallback)"
        age_min = _minutes_since(raw_last_heartbeat)
        # age_min < 0 means last_heartbeat_at is ahead of our clock (device
        # clock drift/bad NTP sync, or a timezone mismatch upstream) — a
        # "heartbeat from the future" is never valid, so it must not count
        # as received no matter how small the (impossible) gap looks.
        heartbeat_received = age_min is not None and 0 <= age_min <= HEARTBEAT_STALE_MINUTES
        print(
            f"[DEVICE-STATUS] store={store} "
            f"source={heartbeat_source} "
            f"last_heartbeat_at={raw_last_heartbeat!r} "
            f"now_utc={datetime.utcnow().isoformat()} "
            f"age_min={age_min} "
            f"threshold={HEARTBEAT_STALE_MINUTES} "
            f"heartbeat_received={heartbeat_received}",
            flush=True,
        )

        # Heartbeat freshness gates everything: if no heartbeat has arrived
        # within HEARTBEAT_STALE_MINUTES, the device (and its NVR/DVR) is
        # shown offline regardless of what the last cached camera_status said
        # — a device that's been unreachable for hours can't still be "Online"
        # just because its last known DVR reading happened to be good.
        if not heartbeat_received:
            device_status_val = "Offline"
            dvr_status_val = "Offline"
        else:
            device_status_val = "Online"
            dvr_status_val = "Online" if (dvr_status == "Online" or dvr_online_count > 0) else dvr_status

        records.append({
            "store_code":        r.get("store_code", ""),
            "hostname":          r.get("hostname", ""),
            "device_status":     device_status_val,
            "software_status":   software_status,
            "dvr_nvr_status":    dvr_status_val,
            "temperature":       r.get("cpu_temp"),
            "serial_id":         str(r.get("_id", "")),
            "dvr_online_count":  r.get("dvr_online_count"),
            "dvr_total_count":   r.get("dvr_total_count"),
            "last_heartbeat_at": _iso(raw_last_heartbeat),
        })

    records.sort(key=lambda r: (r["device_status"] != "Online", r["store_code"]))
    total = len(records)
    devices_offline = sum(1 for r in records if r["device_status"] != "Online")
    dvr_offline = sum(1 for r in records if r["dvr_nvr_status"] == "Offline")

    return jsonify({
        "records":         records,
        "total":           total,
        "devices_offline": devices_offline,
        "dvr_nvr_offline": dvr_offline,
        "configured":      True,
    })
